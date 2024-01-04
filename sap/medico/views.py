from django.http import HttpResponse
from django.forms import modelform_factory
from django.shortcuts import redirect, get_object_or_404
from django.template import loader
from medico.models import Doctor, Paciente, Cita
from medico.forms import MedicoFormulario
from openpyxl.workbook import Workbook
from rest_framework import viewsets, permissions
from medico.serializers import PacienteSerializer, DoctorSerializer, CitaSerializer

MedicoFormulario = modelform_factory(Doctor, exclude=['activo',])

def agregar_medico(request):
    pagina = loader.get_template('agregar.html')
    if request.method == "GET":
        formulario = MedicoFormulario
    elif request.method == "POST":
        formulario = MedicoFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario':formulario}
    return HttpResponse(pagina.render(datos,request))

def ver_medico(request, id):
    doctor = get_object_or_404(Doctor, pk=id)
    datos = {'medico': doctor}
    pagina = loader.get_template('ver.html')
    return HttpResponse(pagina.render(datos, request))

def modificar_medico(request, id):
    pagina = loader.get_template('modificar.html')
    doctor = get_object_or_404(Doctor, pk=id)
    if request.method == 'GET':
        formulario = MedicoFormulario(instance=doctor)
    elif request.method == 'POST':
        formulario = MedicoFormulario(request.POST, instance=doctor)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def eliminar_medico(request, id):
    doctor = get_object_or_404(Doctor, pk=id)
    if doctor:
        doctor.delete()
        return redirect('inicio')

def generar_reporte(request, *args, **kwargs):
    medico = Doctor.objects.order_by('apellido', 'nombre')
    wb = Workbook()
    ws = wb.active
    ws['D1'] = 'REPORTE DE MÉDICOS'
    ws.merge_cells('D1:G1')
    ws['B3'] = 'ID'
    ws['C3'] = 'NOMBRE'
    ws['D3'] = 'APELLIDO'
    ws['E3'] = 'EMAIL'
    ws['F3'] = 'ESPECIALIDAD'
    cont = 4

    for doctor in medico:
        ws.cell(row=cont, column=2).value = doctor.id
        ws.cell(row=cont, column=3).value = doctor.nombre
        ws.cell(row=cont, column=4).value = doctor.apellido
        ws.cell(row=cont, column=5).value = doctor.email
        ws.cell(row=cont, column=6).value = doctor.especialidad
        cont = cont + 1

    nombre_archivo = "ReporteMédicosExcel.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

class PacienteViewSet(viewsets.ModelViewSet):
    queryset = Paciente.objects.all()
    serializer_class = PacienteSerializer
    permission_classes = [permissions.IsAuthenticated]
class DoctorViewSet(viewsets.ModelViewSet):
    queryset = Doctor.objects.all()
    serializer_class = DoctorSerializer
    permission_classes = [permissions.IsAuthenticated]
class CitaViewSet(viewsets.ModelViewSet):
    queryset = Cita.objects.all()
    serializer_class = CitaSerializer
    permission_classes = [permissions.IsAuthenticated]